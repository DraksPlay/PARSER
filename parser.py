import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
import time
import openpyxl
import json
import asyncio


def log(*args, sep=" ", end="\n"):
    print(f"[INFO][{time.strftime('%H:%M:%S')}]: ", *args, sep=sep, end=end)


def get_subjects(driver, try_district=False):
   try:
       regions = {}
       regions_temp = []
       cols_len = len(driver.find_element(By.XPATH, '/html/body/section[5]/div/div').find_elements(By.XPATH, "./*"))
       cols = [driver.find_element(By.XPATH, f'/html/body/section[5]/div/div/div[{i}]') for i in range(1, cols_len+1)]
       for col in cols:
          for elem in col.find_elements(By.TAG_NAME, 'a'):
             regions_temp.append(elem)
       for index, value in enumerate(regions_temp):
          regions[value.text] = {"url": value.get_attribute("href")}
       return regions
   except Exception as exc:
       return False

def get_addresses(driver):
    try:
        url = "https://www.reformagkh.ru/myhouse?page={}&limit=60&view=list&sort=name&order=asc"
        addresses = {}
        li_elems = len(driver.find_element(By.XPATH, "//*[@id='paginatorForm']/div/nav/ul").find_elements(By.XPATH, "./*"))
        numbers_page = driver.find_element(By.XPATH, f"//*[@id='paginatorForm']/div/nav/ul/li[{li_elems}]/a").get_attribute("data-page")
        for number_page in range(1, int(numbers_page) + 1):
            if number_page != 1:
                driver.get(url.format(number_page))
            driver.implicitly_wait(5)
            tbody = driver.find_element(By.XPATH, "//*[@id='myHouseList']/tbody")
            elems = [i.find_element(By.XPATH, "//td[1]/a") for i in tbody.find_elements(By.TAG_NAME, "tr")]
            c = 0
            for address in elems:
                c += 1
                addresses[address.text] = {"url": address.get_attribute("href")}
        return addresses
    except:
        return {}

def parser_data(driver, url):
    try:
        driver.get(url)
        driver.implicitly_wait(2)
        # Объём средств
        redirect_btn = driver.find_element(By.XPATH, '//*[@id="house-card-menu-container"]/a[3]')
        driver.execute_script("arguments[0].click();", redirect_btn)
        driver.implicitly_wait(2)
        data = []
        data.append(driver.find_element(By.XPATH, "//*[@id='list-common']/table/tbody/tr[1]/td[2]").text)
        data.append(driver.find_element(By.XPATH, "//*[@id='list-common']/table/tbody/tr[2]/td[2]").text)
        data.append(driver.find_element(By.XPATH, "//*[@id='list-common']/table/tbody/tr[4]/td[2]").text)
        data.append(driver.find_element(By.XPATH, "//*[@id='list-common']/table/tbody/tr[5]/td[2]").text)
        data.append(driver.find_element(By.XPATH, "//*[@id='list-common']/table/tbody/tr[6]/td[2]").text)
        data.append(driver.find_element(By.XPATH, "//*[@id='list-common']/table/tbody/tr[7]/td[2]").text)
        driver.get(driver.find_element(By.XPATH, "//*[@id='headingservices']/a").get_attribute("href"))
        driver.implicitly_wait(2)
        time.sleep(1)
        for _ in range(6):
            data.append([])
        for index, type_work in enumerate(driver.find_element(By.XPATH, "//*[@id='services-accordion']/div").find_elements(By.XPATH, "./*")):
            button = driver.find_element(By.XPATH, f"//*[@id='services-accordion']/div/ul[{index+1}]/li/a")
            driver.execute_script("arguments[0].click();", button)
            time.sleep(0.5)
            for table in range(1, len(type_work.find_element(By.XPATH, f"//*[@id='services-collapse{index+1}']").find_elements(By.XPATH, "./*"))+1):
                for i in range(1, len(type_work.find_element(By.XPATH, f"//*[@id='services-collapse{index + 1}']/table[{table}]/tbody[2]/tr[1]").find_elements(By.TAG_NAME, "td"))):
                    data[index+6].append(type_work.find_element(By.XPATH, f"//*[@id='services-collapse{index + 1}']/table[{table}]/tbody[2]/tr[1]/td[{i}]").text)

        return data
    except:
        return []

def save_data(data, row, title, sheet, workbook, key):

    sheet.cell(row, 1, data[0])
    sheet.cell(row, 2, data[1])
    sheet.cell(row, 3, data[2])
    col = 4
    for index_elem, value_elem in enumerate(data[3:]):
        if type(value_elem) == list:
            for elem2 in value_elem:
                print(elem2)
                sheet.cell(row, col, elem2)
                col += 1
        else:
            sheet.cell(row, col, value_elem)
            col += 1
    title = key+".xlsx"
    workbook.save(title)

slice = {"district": "Балахнинский муниципальный округ", "selsovet": "Коневский сельсовет"}
async def parser_address(driver, region="all"):
    regions = get_subjects(driver)
    for key, value in regions.items():
        workbook = openpyxl.load_workbook("header.xlsx")
        sheet = workbook.active
        if region == key or region == "all":
            driver.get(value["url"])
            driver.implicitly_wait(2)
            regions[key]["districts"] = get_subjects(driver)
            count = 2
            district_slice = {}
            if slice["district"]:
                for slice_index, slice_value in regions[key]["districts"].items():
                    if slice_index == slice["district"]:
                        district_slice.clear()
                    district_slice[slice_index] = slice_value
                slice["district"] = False
            for key_district, value_disctrict in district_slice.items():
                driver.get(value_disctrict["url"])
                driver.implicitly_wait(2)
                gs = get_subjects(driver, try_district=True)
                regions[key]["districts"][key_district]["selsovet"] = {key_district: {"url": value_disctrict["url"]}} if not gs else gs
                selsovet_slice = {}
                if slice["selsovet"]:
                    for slice_index, slice_value in regions[key]["districts"][key_district]["selsovet"].items():
                        if slice_index == slice["selsovet"]:
                            selsovet_slice.clear()
                        else:
                            selsovet_slice[slice_index] = slice_value
                    slice["selsovet"] = False
                for key_selsovet, value_selsovet in selsovet_slice.items():
                    log(key, key_district, key_selsovet)
                    driver.get(value_selsovet["url"])
                    driver.implicitly_wait(2)
                    ga = get_addresses(driver)
                    regions[key]["districts"][key_district]["selsovet"][key_selsovet]["addresses"] = ga
                    with open("data.json", "w", encoding="utf-8") as file:
                        json.dump(regions, file, indent=4, ensure_ascii=False)
                        file.close()
                    """for key_address, value_address in regions[key]["districts"][key_district]["selsovet"][key_selsovet]["addresses"].items():
                        pa = parser_data(driver, value_address["url"])
                        print("Parsed:", key_address)
                        regions[key]["districts"][key_district]["selsovet"][key_selsovet]["addresses"][key_address]["data"] = pa
                        data = [key_district, key_selsovet, key_address] + pa
                        #threading.Thread(target=save_data, args=[data, count, key]).run()
                        save_data(data, count, key, sheet, workbook, key)
                        count += 1"""




    return regions

async def main():
    try:
        chrome_options = Options()
        #chrome_options.add_argument("--headless")
        region = input("Введите область или all, чтобы парсить все области: ")

        #chrome_options.add_argument("--proxy-server=%s" % hostname + ":" + port)
        driver = webdriver.Chrome(options=chrome_options)
        log("Parser started")
        driver.get("https://www.reformagkh.ru/myhouse?tid=2208161")
        driver.implicitly_wait(5)

        data = asyncio.create_task(parser_address(driver, region=region))

        await data
        while True:
            msg = input("MSG ")
            print(msg)
        import json
        with open("data.json", "w", encoding="utf-8") as file:
            json.dump(data, file, indent=4, ensure_ascii=False)
            file.close()
        log("Data was saved")
        log("Parser was finished!")

        driver.quit()
    except Exception as exc:
        raise


if __name__ == '__main__':
    asyncio.run(main())